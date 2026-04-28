#!/usr/bin/env python3
"""
Solar Edge Roster Reconciliation Tool
--------------------------------------
Usage:
  python solar_reconcile.py --email email.txt --timesheets ./timesheets/
  python solar_reconcile.py --email email.txt --timesheets ./timesheets/ --output report.xlsx

Arguments:
  --email       Path to a .txt file containing the Solar Edge email body (or paste mode if omitted)
  --timesheets  Path to a folder containing timesheet .xlsx files (one sheet per associate)
  --output      (optional) Path to save the Excel reconciliation report (default: reconciliation_report.xlsx)
  --shift       (optional) Shift name for the report header (e.g. "2nd shift")

How the email file should look (copy/paste from the Solar Edge email body):
  Employee Summary for 4/23/2026, 2nd shift, Simcoe Parts Service Inc.
  Employee | Emp ID # | Payable Hours | Training Hours | Bench Hours
  Justin Coffey | T245932 | 9.00 | 0.00 | 0.00
  Paul Crowe | T243960 | 9.00 | 0.00 | 0.00
  ...

Timesheet folder:
  Each .xlsx file should have one sheet per associate, with:
  - Row 8, Col 5 (F): Associate name
  - Row 28, Col 3 (D): Total hours as a timedelta
"""

import argparse
import datetime
import glob
import os
import re
import sys
from difflib import SequenceMatcher

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependencies. Run: pip install pandas openpyxl")
    sys.exit(1)


# ─── Parsing ────────────────────────────────────────────────────────────────

def parse_email_file(path):
    """Parse Solar Edge email body file into list of dicts."""
    with open(path, "r", encoding="utf-8") as f:
        text = f.read()
    return parse_email_text(text)


def parse_email_text(text):
    """Parse raw email text into roster list."""
    employees = []
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    shift_date = None
    shift_name = None

    for line in lines:
        # Try to extract date and shift from header
        m = re.search(r"Employee Summary for\s+(\d+/\d+/\d+),?\s*(.*?),?\s*(?:Simcoe|$)", line, re.IGNORECASE)
        if m:
            shift_date = m.group(1)
            shift_name = m.group(2).strip()
            continue

        # Split by tab or pipe
        parts = re.split(r"\t|\|", line)
        parts = [p.strip() for p in parts if p.strip()]
        if len(parts) < 3:
            continue

        # Skip headers
        if re.match(r"employee|emp\s*id|payable", parts[0], re.IGNORECASE):
            continue

        try:
            hours = float(parts[2])
        except ValueError:
            continue

        employees.append({
            "name": parts[0],
            "emp_id": parts[1] if len(parts) > 1 else "—",
            "roster_hours": hours,
            "training_hours": float(parts[3]) if len(parts) > 3 else 0.0,
            "bench_hours": float(parts[4]) if len(parts) > 4 else 0.0,
        })

    return employees, shift_date, shift_name


def read_timesheets(folder):
    """
    Read all .xlsx files in folder.
    Returns dict: { normalized_name -> {"name": str, "hours": float, "file": str} }
    """
    results = {}
    pattern = os.path.join(folder, "**", "*.xlsx")
    files = glob.glob(pattern, recursive=True)
    if not files:
        # Also try non-recursive
        files = glob.glob(os.path.join(folder, "*.xlsx"))

    for fpath in files:
        try:
            xl = pd.read_excel(fpath, sheet_name=None, header=None)
        except Exception as e:
            print(f"  Warning: could not read {fpath}: {e}")
            continue

        for sheet_name, df in xl.items():
            name = None
            hours = None

            # Name: row 8, col 5 (0-indexed)
            try:
                name_val = df.iloc[8, 5]
                if pd.notna(name_val) and str(name_val).strip():
                    name = str(name_val).strip()
            except (IndexError, KeyError):
                pass

            # Fallback: use sheet name as name
            if not name:
                name = sheet_name.strip()

            # Hours: scan all rows for "Total Hours Worked for Week" label
            # then find the non-zero datetime.time value in that row
            try:
                for i, row in df.iterrows():
                    for j, val in enumerate(row):
                        if "Total Hours Worked" in str(val):
                            # Found the label row — now scan across for a time value
                            for c in range(len(df.columns)):
                                cell_val = df.iloc[i, c]
                                if isinstance(cell_val, datetime.time) and (cell_val.hour > 0 or cell_val.minute > 0):
                                    hours = cell_val.hour + cell_val.minute / 60 + cell_val.second / 3600
                                    break
                                elif isinstance(cell_val, datetime.timedelta):
                                    hours = cell_val.total_seconds() / 3600
                                    break
                                elif isinstance(cell_val, (int, float)) and not pd.isna(cell_val) and cell_val > 0:
                                    hours = float(cell_val)
                                    break
                            break
                    if hours is not None:
                        break
            except Exception:
                pass

            if name:
                key = normalize_name(name)
                results[key] = {
                    "name": name,
                    "hours": hours,
                    "file": os.path.basename(fpath),
                    "sheet": sheet_name,
                }

    return results


# ─── Name matching ──────────────────────────────────────────────────────────

def normalize_name(name):
    """Lowercase, strip punctuation, sort words for fuzzy matching."""
    name = re.sub(r"[^a-zA-Z\s]", "", name).lower().strip()
    return " ".join(sorted(name.split()))


def fuzzy_match(name, candidates, threshold=0.75):
    """Find best fuzzy match from candidates dict keys."""
    key = normalize_name(name)
    best_score = 0
    best_match = None

    for candidate_key, data in candidates.items():
        score = SequenceMatcher(None, key, candidate_key).ratio()
        if score > best_score:
            best_score = score
            best_match = candidate_key

    if best_score >= threshold:
        return candidates[best_match], best_score
    return None, best_score


# ─── Reconciliation ─────────────────────────────────────────────────────────

def reconcile(roster, timesheets):
    """Match roster entries to timesheet data and compute discrepancies."""
    results = []
    for emp in roster:
        ts_data, score = fuzzy_match(emp["name"], timesheets)
        if ts_data and ts_data["hours"] is not None:
            ts_hours = ts_data["hours"]
            diff = ts_hours - emp["roster_hours"]
            status = "MATCH" if abs(diff) < 0.01 else "DISCREPANCY"
            results.append({
                **emp,
                "ts_hours": ts_hours,
                "diff": diff,
                "status": status,
                "ts_file": ts_data["file"],
                "match_score": round(score, 2),
                "matched_name": ts_data["name"],
            })
        else:
            results.append({
                **emp,
                "ts_hours": None,
                "diff": None,
                "status": "MISSING TIMESHEET",
                "ts_file": ts_data["file"] if ts_data else "—",
                "match_score": round(score, 2),
                "matched_name": ts_data["name"] if ts_data else "—",
            })
    return results


# ─── Excel report ────────────────────────────────────────────────────────────

def write_excel_report(results, output_path, shift_date=None, shift_name=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    # Colors
    GREEN = "C6EFCE"
    RED = "FFC7CE"
    AMBER = "FFEB9C"
    HEADER_BG = "1F3864"
    HEADER_FG = "FFFFFF"
    SUBHEADER_BG = "D9E1F2"
    MATCH_FG = "276221"
    DISC_FG = "9C0006"
    MISS_FG = "7D4C00"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell_style(ws, row, col, value, bold=False, bg=None, fg=None, align="left", num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, color=fg or "000000", name="Calibri", size=11)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = border
        if num_fmt:
            c.number_format = num_fmt
        return c

    # Title
    ws.merge_cells("A1:J1")
    title = f"Solar Edge Roster Reconciliation"
    if shift_date:
        title += f"  —  {shift_date}"
    if shift_name:
        title += f"  ({shift_name})"
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, color=HEADER_FG, name="Calibri", size=13)
    c.fill = PatternFill("solid", fgColor=HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Summary row
    total = len(results)
    matched = sum(1 for r in results if r["status"] == "MATCH")
    discs = sum(1 for r in results if r["status"] == "DISCREPANCY")
    missing = sum(1 for r in results if r["status"] == "MISSING TIMESHEET")
    ws.merge_cells("A2:J2")
    summary = f"Total: {total}   |   Matched: {matched}   |   Discrepancies: {discs}   |   Missing timesheets: {missing}"
    c = ws["A2"]
    c.value = summary
    c.font = Font(bold=False, color="1F3864", name="Calibri", size=10)
    c.fill = PatternFill("solid", fgColor=SUBHEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Headers
    headers = ["Employee", "Emp ID", "Roster Hrs", "Timesheet Hrs", "Difference", "Status", "Matched Name", "Match Score", "Timesheet File", "Notes"]
    for col, h in enumerate(headers, 1):
        cell_style(ws, 3, col, h, bold=True, bg=HEADER_BG, fg=HEADER_FG, align="center")
    ws.row_dimensions[3].height = 22

    # Data rows
    for row_i, r in enumerate(results, 4):
        ws.row_dimensions[row_i].height = 18
        status = r["status"]
        if status == "MATCH":
            row_bg = GREEN
            status_fg = MATCH_FG
        elif status == "DISCREPANCY":
            row_bg = RED
            status_fg = DISC_FG
        else:
            row_bg = AMBER
            status_fg = MISS_FG

        cell_style(ws, row_i, 1, r["name"], bg=row_bg)
        cell_style(ws, row_i, 2, r["emp_id"], bg=row_bg, align="center")
        cell_style(ws, row_i, 3, r["roster_hours"], bg=row_bg, align="center", num_fmt="0.00")
        ts_val = r["ts_hours"] if r["ts_hours"] is not None else "—"
        cell_style(ws, row_i, 4, ts_val, bg=row_bg, align="center", num_fmt="0.00" if r["ts_hours"] is not None else None)

        diff_val = r["diff"]
        if diff_val is not None:
            diff_str = f"+{diff_val:.2f}" if diff_val > 0 else f"{diff_val:.2f}"
            if abs(diff_val) < 0.01:
                diff_str = "0.00"
        else:
            diff_str = "—"
        cell_style(ws, row_i, 5, diff_str, bg=row_bg, align="center")
        cell_style(ws, row_i, 6, status, bold=True, bg=row_bg, fg=status_fg, align="center")
        cell_style(ws, row_i, 7, r.get("matched_name", "—"), bg=row_bg)
        cell_style(ws, row_i, 8, r.get("match_score", "—"), bg=row_bg, align="center")
        cell_style(ws, row_i, 9, r.get("ts_file", "—"), bg=row_bg)
        cell_style(ws, row_i, 10, "", bg=row_bg)  # Notes column left blank

    # Column widths
    widths = [28, 12, 13, 15, 12, 22, 28, 13, 30, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(output_path)
    return output_path


# ─── CLI ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Solar Edge Roster Reconciliation Tool")
    parser.add_argument("--email", required=False, help="Path to email body .txt file")
    parser.add_argument("--timesheets", required=True, help="Path to folder containing timesheet .xlsx files")
    parser.add_argument("--output", default="reconciliation_report.xlsx", help="Output Excel report path")
    parser.add_argument("--shift", default=None, help="Shift label for the report")
    args = parser.parse_args()

    print("\n── Solar Edge Reconciliation ──────────────────")

    # Get email content
    if args.email:
        if not os.path.exists(args.email):
            print(f"Error: email file not found: {args.email}")
            sys.exit(1)
        print(f"Reading email from: {args.email}")
        roster, shift_date, shift_name = parse_email_file(args.email)
    else:
        print("Paste the Solar Edge email body below, then press Enter twice + Ctrl+D:")
        lines = []
        try:
            while True:
                line = input()
                lines.append(line)
        except EOFError:
            pass
        roster, shift_date, shift_name = parse_email_text("\n".join(lines))

    if args.shift:
        shift_name = args.shift

    if not roster:
        print("Error: no employees found in email. Check format.")
        sys.exit(1)

    print(f"Found {len(roster)} employees in roster.")

    # Read timesheets
    if not os.path.exists(args.timesheets):
        print(f"Error: timesheets folder not found: {args.timesheets}")
        sys.exit(1)

    print(f"Reading timesheets from: {args.timesheets}")
    timesheets = read_timesheets(args.timesheets)
    print(f"Found {len(timesheets)} timesheet entries.")

    # Reconcile
    results = reconcile(roster, timesheets)

    # Print summary to terminal
    print("\n── Results ────────────────────────────────────")
    matched = [r for r in results if r["status"] == "MATCH"]
    discs = [r for r in results if r["status"] == "DISCREPANCY"]
    missing = [r for r in results if r["status"] == "MISSING TIMESHEET"]

    print(f"  Matched:            {len(matched)}")
    print(f"  Discrepancies:      {len(discs)}")
    print(f"  Missing timesheets: {len(missing)}")

    if discs:
        print("\n── Discrepancies ──────────────────────────────")
        for r in discs:
            diff = r["diff"]
            sign = "+" if diff > 0 else ""
            print(f"  {r['name']:30s}  Roster: {r['roster_hours']:.2f}  |  Timesheet: {r['ts_hours']:.2f}  |  Diff: {sign}{diff:.2f}")

    if missing:
        print("\n── Missing Timesheets ─────────────────────────")
        for r in missing:
            print(f"  {r['name']:30s}  Roster: {r['roster_hours']:.2f}")

    # Write Excel report
    out = write_excel_report(results, args.output, shift_date, shift_name)
    print(f"\n  Report saved: {out}")
    print("──────────────────────────────────────────────\n")


if __name__ == "__main__":
    main()
