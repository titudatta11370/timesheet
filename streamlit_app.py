import streamlit as st
import pandas as pd
import datetime
import re
import io
import os
import tempfile
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Solar Edge Reconciliation",
    page_icon="📋",
    layout="centered"
)

# ─── Auth gate ───────────────────────────────────────────────────────────────

def _check_login():
    """Single-user username/password gate backed by st.secrets.

    Expected secrets (in .streamlit/secrets.toml locally or Streamlit Cloud → Secrets):
        app_username = "yourname"
        app_password = "yourpassword"
    """
    expected_user = st.secrets.get("app_username", None)
    expected_pass = st.secrets.get("app_password", None)

    if not expected_user or not expected_pass:
        st.error(
            "App is not configured. Set `app_username` and `app_password` in "
            "Streamlit secrets (Settings → Secrets on Streamlit Cloud, or "
            "`.streamlit/secrets.toml` locally)."
        )
        st.stop()

    if st.session_state.get("authenticated"):
        return

    st.title("🔒 Sign in")
    st.caption("Solar Edge Roster Reconciliation")
    with st.form("login_form", clear_on_submit=False):
        u = st.text_input("Username")
        p = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Sign in", type="primary", use_container_width=True)
    if submitted:
        if u.strip().lower() == expected_user.strip().lower() and p == expected_pass:
            st.session_state["authenticated"] = True
            st.session_state["user"] = expected_user
            st.rerun()
        else:
            st.error("Invalid username or password.")
    st.stop()

_check_login()

# Sidebar: who's logged in + logout
with st.sidebar:
    st.markdown(f"**Signed in as:** `{st.session_state.get('user', '')}`")
    if st.button("Log out", use_container_width=True):
        for k in ("authenticated", "user"):
            st.session_state.pop(k, None)
        st.rerun()

st.title("📋 Solar Edge Roster Reconciliation")
st.markdown("Upload your Solar Edge email and timesheet Excel files to generate a reconciliation report.")

# ─── Email extraction ────────────────────────────────────────────────────────

def extract_email_text(uploaded_file):
    filename = uploaded_file.name.lower()
    if filename.endswith(".msg"):
        try:
            import extract_msg
            with tempfile.NamedTemporaryFile(delete=False, suffix=".msg") as tmp:
                tmp.write(uploaded_file.read())
                tmp_path = tmp.name
            msg = extract_msg.openMsg(tmp_path)
            text = msg.body or ""
            msg.close()
            os.unlink(tmp_path)
            return text
        except ImportError:
            st.error("Missing library: extract-msg. Contact the app admin.")
            return ""
        except Exception as e:
            st.error(f"Could not read .msg file: {e}")
            return ""
    else:
        return uploaded_file.read().decode("utf-8", errors="ignore")

# ─── Helper functions ────────────────────────────────────────────────────────

def normalize_name(name):
    name = re.sub(r"[^a-zA-Z\s]", "", name).lower().strip()
    return " ".join(sorted(name.split()))

def fuzzy_match(name, candidates, threshold=0.75):
    key = normalize_name(name)
    best_score = 0
    best_match = None
    for candidate_key, data in candidates.items():
        score = SequenceMatcher(None, key, candidate_key).ratio()
        if score > best_score:
            best_score = score
            best_match = candidate_key
    if best_score >= threshold:
        return candidates[best_match], best_score
    return None, best_score

def parse_email_text(text):
    employees = []
    shift_date = None
    shift_name = None
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    for line in lines:
        m = re.search(r"Employee Summary for\s+(\d+/\d+/\d+),?\s*(.*?)(?:,\s*Simcoe|$)", line, re.IGNORECASE)
        if m:
            shift_date = m.group(1)
            shift_name = m.group(2).strip()
            continue
        parts = re.split(r"\t|\|", line)
        parts = [p.strip() for p in parts if p.strip()]
        if len(parts) < 3:
            continue
        if re.match(r"employee|emp\s*id|payable", parts[0], re.IGNORECASE):
            continue
        try:
            hours = float(parts[2])
        except ValueError:
            continue
        employees.append({
            "name": parts[0],
            "emp_id": parts[1] if len(parts) > 1 else "—",
            "roster_hours": hours,
        })
    return employees, shift_date, shift_name

def time_to_hours(val):
    """Convert datetime.time or timedelta to float hours."""
    if isinstance(val, datetime.time):
        return val.hour + val.minute / 60 + val.second / 3600
    elif isinstance(val, datetime.timedelta):
        return val.total_seconds() / 3600
    elif isinstance(val, (int, float)) and not pd.isna(val) and val > 0:
        return float(val)
    return None

def read_timesheets_from_files(uploaded_files):
    """
    Read timesheet files and extract per-person daily hours by date plus weekly total.
    Aggregates across multiple sheets/files when the same associate appears in
    multiple plants (sums daily hours per date and sums weekly totals).
    """
    results = {}
    for uploaded_file in uploaded_files:
        try:
            xl = pd.read_excel(uploaded_file, sheet_name=None, header=None)
        except Exception as e:
            st.warning(f"Could not read {uploaded_file.name}: {e}")
            continue
        for sheet_name, df in xl.items():
            name = None
            weekly_hours = None

            try:
                name_val = df.iloc[8, 5]
                if pd.notna(name_val) and str(name_val).strip():
                    name = str(name_val).strip()
            except (IndexError, KeyError):
                pass
            if not name:
                name = sheet_name.strip()

            dates_row = None
            daily_hours_row = None
            try:
                for i, row in df.iterrows():
                    for j, val in enumerate(row):
                        if isinstance(val, (datetime.datetime, pd.Timestamp)):
                            if dates_row is None:
                                dates_row = i
                        sval = str(val)
                        if "Total Daily Hours" in sval:
                            daily_hours_row = i
                        if "Total Hours Worked" in sval:
                            for c in range(len(df.columns)):
                                h = time_to_hours(df.iloc[i, c])
                                if h and h > 0:
                                    weekly_hours = h
                                    break
            except Exception:
                pass

            daily_by_date = {}
            if dates_row is not None and daily_hours_row is not None:
                for c in range(len(df.columns)):
                    cell_val = df.iloc[dates_row, c]
                    cell_date = None
                    if isinstance(cell_val, (datetime.datetime, pd.Timestamp)):
                        cell_date = cell_val.date() if hasattr(cell_val, 'date') else cell_val
                    if cell_date:
                        h = time_to_hours(df.iloc[daily_hours_row, c])
                        if h is not None:
                            daily_by_date[cell_date] = h

            if not name:
                continue

            key = normalize_name(name)
            source = f"{uploaded_file.name} / {sheet_name}"
            if key in results:
                entry = results[key]
                for d, h in daily_by_date.items():
                    entry["daily_by_date"][d] = entry["daily_by_date"].get(d, 0) + h
                if weekly_hours is not None:
                    entry["weekly_hours"] = (entry["weekly_hours"] or 0) + weekly_hours
                entry["sources"].append(source)
            else:
                results[key] = {
                    "name": name,
                    "weekly_hours": weekly_hours,
                    "daily_by_date": daily_by_date,
                    "sources": [source],
                }
    return results

def reconcile(roster, timesheets, target_date=None):
    results = []
    for emp in roster:
        ts_data, score = fuzzy_match(emp["name"], timesheets)
        if ts_data:
            ts_daily = ts_data["daily_by_date"].get(target_date) if target_date else None
            ts_weekly = ts_data.get("weekly_hours")
            ts_hours = ts_daily if ts_daily is not None else ts_weekly
            hours_source = "Daily" if ts_daily is not None else "Weekly total"

            if ts_hours is not None:
                diff = ts_hours - emp["roster_hours"]
                status = "MATCH" if abs(diff) < 0.01 else "DISCREPANCY"
            else:
                diff = None
                status = "MISSING TIMESHEET"

            results.append({
                **emp,
                "ts_hours": ts_hours,
                "ts_daily": ts_daily,
                "ts_weekly": ts_weekly,
                "hours_source": hours_source,
                "diff": diff,
                "status": status,
                "ts_file": "; ".join(ts_data["sources"]),
                "match_score": round(score, 2),
                "matched_name": ts_data["name"],
            })
        else:
            results.append({
                **emp,
                "ts_hours": None,
                "ts_daily": None,
                "ts_weekly": None,
                "hours_source": "—",
                "diff": None,
                "status": "MISSING TIMESHEET",
                "ts_file": "—",
                "match_score": round(score, 2),
                "matched_name": "—",
            })
    return results

def build_excel_report(results, shift_date=None, shift_name=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    GREEN = "C6EFCE"; RED = "FFC7CE"; AMBER = "FFEB9C"
    HEADER_BG = "1F3864"; HEADER_FG = "FFFFFF"
    SUBHEADER_BG = "D9E1F2"
    MATCH_FG = "276221"; DISC_FG = "9C0006"; MISS_FG = "7D4C00"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cs(row, col, value, bold=False, bg=None, fg=None, align="left", num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, color=fg or "000000", name="Calibri", size=11)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = border
        if num_fmt:
            c.number_format = num_fmt
        return c

    ws.merge_cells("A1:M1")
    title = "Solar Edge Roster Reconciliation"
    if shift_date:
        title += f"  —  {shift_date}"
    if shift_name:
        title += f"  ({shift_name})"
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, color=HEADER_FG, name="Calibri", size=13)
    c.fill = PatternFill("solid", fgColor=HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    total = len(results)
    matched = sum(1 for r in results if r["status"] == "MATCH")
    discs = sum(1 for r in results if r["status"] == "DISCREPANCY")
    missing = sum(1 for r in results if r["status"] == "MISSING TIMESHEET")

    ws.merge_cells("A2:M2")
    summary = f"Total: {total}   |   Matched: {matched}   |   Discrepancies: {discs}   |   Missing: {missing}"
    c = ws["A2"]
    c.value = summary
    c.font = Font(color="1F3864", name="Calibri", size=10)
    c.fill = PatternFill("solid", fgColor=SUBHEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Date", "Shift", "Employee", "Emp ID", "Roster Hrs\n(Email Date)", "Timesheet Hrs\n(Same Day)", "Weekly Total\n(Timesheet)", "Difference", "Status", "Hours Source", "Matched Name", "Timesheet File(s)", "Notes"]
    for col, h in enumerate(headers, 1):
        cs(3, col, h, bold=True, bg=HEADER_BG, fg=HEADER_FG, align="center")
    ws.row_dimensions[3].height = 30

    for row_i, r in enumerate(results, 4):
        ws.row_dimensions[row_i].height = 18
        status = r["status"]
        row_bg = GREEN if status == "MATCH" else RED if status == "DISCREPANCY" else AMBER
        status_fg = MATCH_FG if status == "MATCH" else DISC_FG if status == "DISCREPANCY" else MISS_FG

        cs(row_i, 1, r.get("shift_date", "—"), bg=row_bg, align="center")
        cs(row_i, 2, r.get("shift_name", "—"), bg=row_bg, align="center")
        cs(row_i, 3, r["name"], bg=row_bg)
        cs(row_i, 4, r["emp_id"], bg=row_bg, align="center")
        cs(row_i, 5, r["roster_hours"], bg=row_bg, align="center", num_fmt="0.00")
        daily_val = r.get("ts_daily")
        cs(row_i, 6, daily_val if daily_val is not None else "—", bg=row_bg, align="center", num_fmt="0.00" if daily_val is not None else None)
        weekly_val = r.get("ts_weekly")
        cs(row_i, 7, weekly_val if weekly_val is not None else "—", bg=row_bg, align="center", num_fmt="0.00" if weekly_val is not None else None)
        diff_val = r["diff"]
        if diff_val is not None:
            diff_str = f"+{diff_val:.2f}" if diff_val > 0 else ("0.00" if abs(diff_val) < 0.01 else f"{diff_val:.2f}")
        else:
            diff_str = "—"
        cs(row_i, 8, diff_str, bg=row_bg, align="center")
        cs(row_i, 9, status, bold=True, bg=row_bg, fg=status_fg, align="center")
        cs(row_i, 10, r.get("hours_source", "—"), bg=row_bg, align="center")
        cs(row_i, 11, r.get("matched_name", "—"), bg=row_bg)
        cs(row_i, 12, r.get("ts_file", "—"), bg=row_bg)
        cs(row_i, 13, "", bg=row_bg)

    widths = [12, 18, 28, 12, 16, 16, 16, 12, 22, 14, 28, 36, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─── UI ──────────────────────────────────────────────────────────────────────

st.markdown("---")

col1, col2 = st.columns(2)

with col1:
    st.subheader("📧 Step 1 — Solar Edge Emails")
    email_files = st.file_uploader(
        "Upload one or more emails as .msg or .txt",
        type=["txt", "msg"],
        accept_multiple_files=True,
        help="Upload one email per shift/day. Upload all 7 days for a full weekly report."
    )

with col2:
    st.subheader("📊 Step 2 — Timesheet Excel Files")
    timesheet_files = st.file_uploader(
        "Upload one or more timesheet .xlsx files",
        type=["xlsx"],
        accept_multiple_files=True,
        help="Upload all the timesheet Excel files for this week"
    )

st.markdown("---")

if st.button("▶ Run Reconciliation", type="primary", use_container_width=True):
    if not email_files:
        st.error("Please upload at least one Solar Edge email file.")
    elif not timesheet_files:
        st.error("Please upload at least one timesheet Excel file.")
    else:
        with st.spinner("Running reconciliation..."):
            # Read timesheets once — aggregates across multi-plant sheets per associate
            timesheets = read_timesheets_from_files(timesheet_files)

            results = []
            parsed_dates = []
            for email_file in email_files:
                email_text = extract_email_text(email_file)
                roster, shift_date, shift_name = parse_email_text(email_text)

                if not roster:
                    st.warning(f"Could not find employees in {email_file.name} — skipping.")
                    continue

                target_date = None
                if shift_date:
                    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
                        try:
                            target_date = datetime.datetime.strptime(shift_date.strip(), fmt).date()
                            break
                        except ValueError:
                            continue

                parsed_dates.append((email_file.name, shift_date, shift_name, target_date, len(roster)))

                day_results = reconcile(roster, timesheets, target_date)
                for r in day_results:
                    r["shift_date"] = shift_date or "—"
                    r["shift_name"] = shift_name or "—"
                results.extend(day_results)

            if parsed_dates:
                with st.expander(f"📅 Processed {len(parsed_dates)} email(s)"):
                    st.dataframe(pd.DataFrame(parsed_dates, columns=["File", "Date", "Shift", "Parsed Date", "Roster Size"]), hide_index=True, use_container_width=True)

            # Use the first email's date/shift as the report header anchor
            shift_date = parsed_dates[0][1] if parsed_dates else None
            shift_name = parsed_dates[0][2] if parsed_dates else None

            if not results:
                st.error("No reconciliation rows produced. Check email file formats.")
            else:

                # Summary metrics
                matched = [r for r in results if r["status"] == "MATCH"]
                discs = [r for r in results if r["status"] == "DISCREPANCY"]
                missing = [r for r in results if r["status"] == "MISSING TIMESHEET"]

                st.success("✅ Reconciliation complete!")

                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Total Employees", len(results))
                m2.metric("✅ Matched", len(matched))
                m3.metric("⚠️ Discrepancies", len(discs))
                m4.metric("❌ Missing Timesheet", len(missing))

                st.markdown("---")

                # Results table
                if discs:
                    st.subheader("⚠️ Discrepancies")
                    disc_df = pd.DataFrame([{
                        "Date": r.get("shift_date", "—"),
                        "Employee": r["name"],
                        "Emp ID": r["emp_id"],
                        "Roster Hrs (Email Date)": r["roster_hours"],
                        "Timesheet Hrs (Same Day)": r.get("ts_daily") if r.get("ts_daily") is not None else "—",
                        "Weekly Total": r.get("ts_weekly") if r.get("ts_weekly") is not None else "—",
                        "Difference": f"+{r['diff']:.2f}" if r["diff"] > 0 else f"{r['diff']:.2f}",
                        "Source": r.get("hours_source", "—")
                    } for r in discs])
                    st.dataframe(disc_df, use_container_width=True, hide_index=True)

                if missing:
                    st.subheader("❌ Missing Timesheets")
                    miss_df = pd.DataFrame([{
                        "Date": r.get("shift_date", "—"),
                        "Employee": r["name"],
                        "Emp ID": r["emp_id"],
                        "Roster Hrs": r["roster_hours"],
                    } for r in missing])
                    st.dataframe(miss_df, use_container_width=True, hide_index=True)

                if matched:
                    with st.expander(f"✅ View {len(matched)} matched employees"):
                        match_df = pd.DataFrame([{
                            "Date": r.get("shift_date", "—"),
                            "Employee": r["name"],
                            "Emp ID": r["emp_id"],
                            "Roster Hrs": r["roster_hours"],
                            "Timesheet Hrs (Day)": r.get("ts_daily") if r.get("ts_daily") is not None else "—",
                            "Weekly Total": r.get("ts_weekly") if r.get("ts_weekly") is not None else "—",
                            "Source": r.get("hours_source", "—")
                        } for r in matched])
                        st.dataframe(match_df, use_container_width=True, hide_index=True)

                # Download report
                st.markdown("---")
                report_date = shift_date.replace("/", "-") if shift_date else datetime.date.today().strftime("%Y-%m-%d")
                excel_report = build_excel_report(results, shift_date, shift_name)
                st.download_button(
                    label="⬇️ Download Excel Report",
                    data=excel_report,
                    file_name=f"reconciliation_{report_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )